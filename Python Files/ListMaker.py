###################################################################################################
# Name        : ListMaker.py
# Author(s)   : Chris Lloyd, Andrew Southwick
# Description : A program to create lists
# Github Link : https://github.com/Clloyd3267/List-Maker/
###################################################################################################

# External Imports
from pathlib import Path # Used for file manipulation
import openpyxl # For reading in verses
from operator import itemgetter
import xlsxwriter # Used to write quizzes to excel files
import time # Used to time exception speed

class ListMaker:
    """
        A class to create different useful lists.

        Attributes:
            debug(bool): A debug variable to enable / disable debug outputs.
            allVerses(array of verse objects) A variable to store all of the verses.
            concordance(dictionary of word objects) A variable to store the concordance.
            uniqueWords(dictionary of word objects) A variable to store all of the unique words.
            twoWordPhrases(dictionary of word objects) A variable to store all of the two word phrases.
            threeWordPhrases(dictionary of word objects) A variable to store all of the three word phrases.
            cvrPhrases(dictionary of word objects) A variable to store all of the cvr phrases.
            crPhrases(dictionary of word objects) A variable to store all of the cr phrases.
            ftvs(array of verse objects) A variable to store all of the ftvs.
            fts(array of partial verse objects) A variable to store all of the fts.
        """

    def __init__(self):
        """
        The constructor for class ListMaker.
        """

        self.debug = "Off"
        self.allVerses = []
        self.concordance = {}
        self.uniqueWords = {}
        self.twoWordPhrases = {}
        self.threeWordPhrases = {}
        self.cvrPhrases = {}
        self.crPhrases = {}
        self.ftvs = []
        self.fts = []

    ####################################################################################################################
    # Main Funcs
    ####################################################################################################################
    def importVerses(self, versesFileName = "Verses.xlsx"):
        """
        Function to import verses from excel file.

        Debug Code (All Verses): "A" or "a" or "On"
        Debug Code (All Verses Split): "W" or "w" or "On"

        Parameters:
            versesFileName (str): The input filename for verse list.

        Returns:
            (0): No errors, (Anything else): Errors.
        """

        # Create the path for Verse file
        dataFilePath = Path("../Data Files/")  # Path where datafiles are stored

        if versesFileName == "Verses.xlsx":
            versesFilePath = dataFilePath / versesFileName
        else:
            versesFilePath = versesFileName

        # Try opening the verses file
        try:
            book = openpyxl.load_workbook(versesFilePath)
        except IOError:
            return "Error => Verses file does not exist!!!"

        sheet = book.worksheets[0]  # Open the first sheet

        # Read in and parse all verses
        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 4):
            # Check to make sure verse is valid
            verse = []
            valid = False

            for cell in row:
                if not cell.value:
                    verse.append("")
                else:
                    verse.append(str(cell.value).strip())
                    valid = True
            # CDL=> Number error codes!!
            if not valid:
                continue
            if not verse[0]:
                return "Error => No Book!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]
            if not verse[1]:
                return "Error => No Chapter!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]
            if not verse[2]:
                return "Error => No Verse Number!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]
            if not verse[3]:
                return "Error => No Verse!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]

            # Replace special characters
            # verse[3] = verse[3].replace("“", "\"")
            # verse[3] = verse[3].replace("”", "\"")
            # verse[3] = verse[3].replace("‘", "\'")
            # verse[3] = verse[3].replace("’", "\'")
            # verse[3] = verse[3].replace("—", "-") CDL=> Remove later

            # Split verse
            verse.append(self.splitVerse(verse[3]))

            # Add verse to list of all verses
            self.allVerses.append(verse)

        # Print All Verses if debug enabled
        if self.debug != "Off" and ("A" in self.debug or "a" in self.debug or self.debug == "On"):
            print("")
            print("=== All Verses (" + str(len(self.allVerses)) + ") ===")
            for verse in self.allVerses:
                print(verse[0] + " " + verse[1] + ":" + verse[2] + " - " + verse[3])

        # Print All Verses Split if debug enabled
        if self.debug != "Off" and ("W" in self.debug or "w" in self.debug or self.debug == "On"):
            print("")
            print("=== All Verses Split (" + str(len(self.allVerses)) + ") ===")
            for verse in self.allVerses:
                print(verse[0] + " " + verse[1] + ":" + verse[2] + " - ", end = "")
                for word in verse[4]:
                    print("(" + word[0] + ")", end = "")
                print("")

        return 0 # Return with no errors

    def createConcordance(self):
        """
        Function to create list of occurrences of words.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "C" or "c" or "On"
        """

        for verse in self.allVerses:
            for i, word in enumerate(verse[4]):
                newVerseText = verse[3][0:word[1]] + "◆" + verse[3][word[1] + len(word[0]):]
                word = str(word[0]).upper()

                if word in self.concordance:
                    self.concordance[word][1].append([verse[0], verse[1], verse[2], newVerseText])
                    self.concordance[word][0] += 1
                else:
                    self.concordance[word] = [1, [[verse[0], verse[1], verse[2], newVerseText]]]

        # Print Concordance if debug enabled
        if self.debug != "Off" and ("C" in self.debug or "c" in self.debug or self.debug == "On"):
            print("")
            print("=== Concordance (" + str(len(self.concordance)) + ") ===")
            for word, value in sorted(self.concordance.items()):
                print(word + " (" + str(self.concordance[word][0]) + ")")
                for occurrence in self.concordance[word][1]:
                    print(" -> " + occurrence[0] + " " + occurrence[1] + ":" + occurrence[2] + " " + str(occurrence[3]))

        return 0 # Return with no errors

    def createUniqueWords(self):
        """
        Function to create list of all Unique Words.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "U" or "u" or "On"
        """

        for word, value in sorted(self.concordance.items()):
            firstOccurence = value[1][0][0:3]
            uniqueWord = True
            for occurence in value[1]:
                if occurence[0:3] != firstOccurence:
                    uniqueWord = False
            if uniqueWord:
                self.uniqueWords[word] = value[1][0][0:3]

        # Print Unique Words if debug enabled
        if self.debug != "Off" and ("U" in self.debug or "u" in self.debug or self.debug == "On"):
            print("")
            print("=== Unique Words (" + str(len(self.uniqueWords)) + ") ===")
            for word, value in sorted(self.uniqueWords.items()):
                print(word + " - " + value[0] + " " + value[1] + ":" + value[2])

        return 0  # Return with no errors

    def createTwoWordPhrases(self):
        """
        Function to create list of two word phrases.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "2" or "On"
        """

        notUniquePhrases = {}
        for verse in self.allVerses:
            i = 0
            while i != len(verse[4]) - 1:
                twoWordPhrase = (str(verse[4][i][0]) + " " + str(verse[4][i + 1][0])).upper()
                if twoWordPhrase not in notUniquePhrases:
                    if twoWordPhrase in self.twoWordPhrases:
                        if self.twoWordPhrases[twoWordPhrase] != verse[0:3]:
                            del self.twoWordPhrases[twoWordPhrase]
                            notUniquePhrases[twoWordPhrase] = verse[0:3]
                    else:
                        self.twoWordPhrases[twoWordPhrase] = verse[0:3]
                i += 1

        # Print Two Word Phrases if debug enabled
        if self.debug != "Off" and ("2" in self.debug or self.debug == "On"):
            print("")
            print("=== Two Word Phrases (" + str(len(self.twoWordPhrases)) + ") ===")
            for phrase, verse in sorted(self.twoWordPhrases.items()):
                print(phrase + " - " + verse[0] + " " + verse[1] + ":" + verse[2])

        return 0  # Return with no errors

    def createThreeWordPhrases(self):
        """
        Function to create list of three word phrases.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "3" or "On"
        """

        notUniquePhrases = {}
        for verse in self.allVerses:
            i = 0
            while i != len(verse[4]) - 2:
                threeWordPhrase = (str(verse[4][i][0]) + " " + str(verse[4][i + 1][0]) + " " +
                                   str(verse[4][i + 2][0])).upper()
                if threeWordPhrase not in notUniquePhrases:
                    if threeWordPhrase in self.threeWordPhrases:
                        if self.threeWordPhrases[threeWordPhrase] != verse[0:3]:
                            del self.threeWordPhrases[threeWordPhrase]
                            notUniquePhrases[threeWordPhrase] = verse[0:3]
                    else:
                        self.threeWordPhrases[threeWordPhrase] = verse[0:3]
                i += 1

        # Print Three Word Phrases if debug enabled
        if self.debug != "Off" and ("3" in self.debug or self.debug == "On"):
            print("")
            print("=== Three Word Phrases (" + str(len(self.threeWordPhrases)) + ") ===")
            for phrase, verse in sorted(self.threeWordPhrases.items()):
                print(phrase + " - " + verse[0] + " " + verse[1] + ":" + verse[2])

        return 0  # Return with no errors

    def createCvrPhrases(self):
        """
        Function to create list of CVR phrases. CDL=> Implement this function.
`
        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "V" or "v" or "On"
        """

        # # Print CVR Phrases if debug enabled
        # if "V" in self.debug or "v" in self.debug or self.debug == "On":
        #     print("")
        #     print("=== CVR Phrases (" + str(len(self.cvrPhrases)) + ") ===")
        #     for phrase, verse in sorted(self.cvrPhrases.items()):

        return 0  # Return with no errors

    def createCrPhrases(self):
        """
        Function to create list of CR phrases. CDL=> Implement this function.

        Returns:
            (0): No errors, (Anything else): Errors.
        """

        return 0  # Return with no errors

    def createFtvs(self):
        """
        Function to create list of first five words of all verses.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "F" or "f" or "On"
        """

        # Add all verses to FTV list
        for verse in self.allVerses:
            firstFiveWords = []
            i = 0
            for i, word in enumerate(verse[4]):
                if i > 4:
                    i -= 1
                    break
                else:
                    firstFiveWords.append(word[0].upper())
            self.ftvs.append(["", firstFiveWords, i, verse[0], verse[1], verse[2], verse[3], verse[4]])

        # Sort the FTV list alphabetically
        self.ftvs = sorted(self.ftvs, key = itemgetter(1, 3, 4, 5))

        currentLine = 0
        while currentLine != len(self.ftvs):
            uniqueNumber = None

            # Get the unique number based on previous verse
            if currentLine != 0:
                word = 0
                while word != 5 and word <= self.ftvs[currentLine][2] and word <= self.ftvs[currentLine - 1][2] \
                    and self.ftvs[currentLine][1][word] == self.ftvs[currentLine - 1][1][word]:
                    word += 1
                uniqueNumber = word

            # Get the unique number based on next verse
            if currentLine != len(self.ftvs) - 1:
                word = 0
                while word != 5 and word <= self.ftvs[currentLine][2] and word <= self.ftvs[currentLine + 1][2] \
                    and self.ftvs[currentLine][1][word] == self.ftvs[currentLine + 1][1][word]:
                    word += 1
                if uniqueNumber == None or uniqueNumber < word:
                    uniqueNumber = word

            # Make the first five words with unique marker
            if uniqueNumber >= len(self.ftvs[currentLine][1]): # If verse not unique after 5 words
                end = self.ftvs[currentLine][7][self.ftvs[currentLine][2]][1]
                while end != len(self.ftvs[currentLine][6]) and self.ftvs[currentLine][6][end] != " ":
                    end += 1
                self.ftvs[currentLine][0] = "||" + self.ftvs[currentLine][6][0:end]
            else:
                mid = self.ftvs[currentLine][7][uniqueNumber][1] + len(self.ftvs[currentLine][7][uniqueNumber][0])
                while mid != len(self.ftvs[currentLine][6]) and self.ftvs[currentLine][6][mid] != " ":
                    mid += 1
                beforeMarker = self.ftvs[currentLine][6][0:mid]

                end = self.ftvs[currentLine][7][self.ftvs[currentLine][2]][1]
                while end != len(self.ftvs[currentLine][6]) and self.ftvs[currentLine][6][end] != " ":
                    end += 1
                afterMarker = self.ftvs[currentLine][6][mid:end]
                self.ftvs[currentLine][0] = beforeMarker + "/" + afterMarker

            currentLine += 1 # Go to next line

        # Print FTVs if debug enabled
        if self.debug != "Off" and ("F" in self.debug or "f" in self.debug or self.debug == "On"):
            print("")
            print("=== FTV Verses (" + str(len(self.ftvs)) + ") ===")
            for verse in self.ftvs:
                print(verse[0], end = "")
                for i in range(45 - len(verse[0])):
                    print("_", end = "")
                print(verse[3] + " " + verse[4] + ":" + verse[5])

        return 0  # Return with no errors

    def createFts(self):
        """
        Function to create list of five words of all valid Fts.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code: "T" or "t" or "On"
        """

        # Parse verses that have valid FTs in them
        tempFts = []
        for verse in self.allVerses:
            for ftQualifier in [" “", " ‘", ". ", "? ", "! ", "; "]:
                if verse[3].find(ftQualifier) != -1:
                    verseText = verse[3][verse[3].find(ftQualifier) + 2:]
                    tempFts.append([verse[0], verse[1], verse[2], verseText, self.splitVerse(verseText)])

        # Add all verses to FT list
        for verse in tempFts:
            firstFiveWords = []
            i = 0
            for i, word in enumerate(verse[4]):
                if i > 4:
                    i -= 1
                    break
                else:
                    firstFiveWords.append(word[0].upper())
            self.fts.append(["", firstFiveWords, i, verse[0], verse[1], verse[2], verse[3], verse[4]])

        # Sort the FT list alphabetically
        self.fts = sorted(self.fts, key = itemgetter(1, 3, 4, 5))

        currentLine = 0
        while currentLine != len(self.fts):
            uniqueNumber = None

            # Get the unique number based on previous verse
            if currentLine != 0:
                word = 0
                while word != 5 and word <= self.fts[currentLine][2] and word <= self.fts[currentLine - 1][2] \
                    and self.fts[currentLine][1][word] == self.fts[currentLine - 1][1][word]:
                    word += 1
                uniqueNumber = word

            # Get the unique number based on next verse
            if currentLine != len(self.fts) - 1:
                word = 0
                while word != 5 and word <= self.fts[currentLine][2] and word <= self.fts[currentLine + 1][2] \
                    and self.fts[currentLine][1][word] == self.fts[currentLine + 1][1][word]:
                    word += 1
                if uniqueNumber == None or uniqueNumber < word:
                    uniqueNumber = word

            # Make the first five words with unique marker
            if uniqueNumber >= len(self.fts[currentLine][1]):  # If verse not unique after 5 words
                end = self.fts[currentLine][7][self.fts[currentLine][2]][1]
                while end != len(self.fts[currentLine][6]) and self.fts[currentLine][6][end] != " ":
                    end += 1
                self.fts[currentLine][0] = "||" + self.fts[currentLine][6][0:end]
            else:
                mid = self.fts[currentLine][7][uniqueNumber][1] + len(self.fts[currentLine][7][uniqueNumber][0])
                while mid != len(self.fts[currentLine][6]) and self.fts[currentLine][6][mid] != " ":
                    mid += 1
                beforeMarker = self.fts[currentLine][6][0:mid]

                end = self.fts[currentLine][7][self.fts[currentLine][2]][1]
                while end != len(self.fts[currentLine][6]) and self.fts[currentLine][6][end] != " ":
                    end += 1
                afterMarker = self.fts[currentLine][6][mid:end]
                self.fts[currentLine][0] = beforeMarker + "/" + afterMarker

            currentLine += 1  # Go to next line

        # Print FTs if debug enabled
        if "T" in self.debug or "t" in self.debug or self.debug == "On":
            print("")
            print("=== FT Verses (" + str(len(self.fts)) + ") ===")
            for verse in self.fts:
                print(verse[0], end = "")
                for i in range(45 - len(verse[0])):
                    print("_", end = "")
                print(verse[3] + " " + verse[4] + ":" + verse[5])

        return 0  # Return with no errors

    def exportLists(self, outputFilename = "Lists.xlsx"):
        """
        Function to export lists.

        Parameters:
            outputFilename(str): The output filename, defaults to "Lists.xlsx".

        Returns:
            (0): No errors, (Anything else): Errors.
        """

        # Create the output file
        if outputFilename == "Lists.xlsx":
            date = time.strftime("%Y_%m_%d")
            fileName = Path("../" + date + "_Lists.xlsx")
            workbook = xlsxwriter.Workbook(fileName)
        else:
            workbook = xlsxwriter.Workbook(outputFilename)

        # Set cell formats
        bold = workbook.add_format({'bold': 1})

        ################################################################################################################
        # Add About worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("About")
        worksheet.set_zoom(175)

        # Add Header cell format
        header = workbook.add_format({'font_size': 18, 'bold': 1})

        # Add actual data
        worksheet.write("A1", "C&MA Bible Quizzing List Maker (V01) by Chris Lloyd.", header)
        currentDate = time.strftime("%Y/%m/%d")
        currentTime = time.strftime("%H:%M")
        worksheet.write("A2", ("Generated on " + currentDate + " at " + currentTime + "."))
        worksheet.write("A3", "For more details: https://github.com/Clloyd3267/List-Maker")
        worksheet.write("A4", "Email Chris Lloyd with any questions, comments, or bugs: Legoman3267@gmail.com")

        ################################################################################################################
        # Add All Verses worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("All Verses")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Verse Text", bold)

        # Add actual data
        i = 2
        for verse in self.allVerses:
            worksheet.write("A" + str(i), verse[0])
            worksheet.write("B" + str(i), verse[1])
            worksheet.write("C" + str(i), verse[2])
            worksheet.write_rich_string("D" + str(i), *self.boldUniqueWords(verse[3], bold))
            i += 1

        ################################################################################################################
        # All Verses Split worksheet (For Searching On)
        ################################################################################################################
        worksheet = workbook.add_worksheet("All Verses Split")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Verse Text Split", bold)

        # Add actual data
        i = 2
        for verse in self.allVerses:
            worksheet.write("A" + str(i), verse[0])
            worksheet.write("B" + str(i), verse[1])
            worksheet.write("C" + str(i), verse[2])
            verseText = ""
            for word in verse[4]:
                verseText += word[0] + " "
            worksheet.write("D" + str(i), verseText.strip())
            i += 1

        ################################################################################################################
        # Add Concordance worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("Concordance")

        # Add headers
        worksheet.write("A1", "Word", bold)
        worksheet.write("B1", "Book", bold)
        worksheet.write("C1", "Chapter", bold)
        worksheet.write("D1", "Verse", bold)
        worksheet.write("E1", "Occurrence", bold)

        # Add actual data
        i = 2
        for word, value in sorted(self.concordance.items()):
            worksheet.write("A" + str(i), word + " (" + str(self.concordance[word][0]) + ")", bold)
            i += 1
            for occurrence in self.concordance[word][1]:
                worksheet.write("A" + str(i), word)
                worksheet.write("B" + str(i), occurrence[0])
                worksheet.write("C" + str(i), occurrence[1])
                worksheet.write("D" + str(i), occurrence[2])
                worksheet.write_rich_string("E" + str(i), *self.boldUniqueWords(occurrence[3], bold))
                i += 1

        ################################################################################################################
        # Add Unique Words worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("Unique Words")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Unique Word", bold)

        # Add actual data
        i = 2
        for word, value in sorted(self.uniqueWords.items()):
            worksheet.write("A" + str(i), value[0])
            worksheet.write("B" + str(i), value[1])
            worksheet.write("C" + str(i), value[2])
            worksheet.write("D" + str(i), word)
            i += 1

        ################################################################################################################
        # Add Two Word Phrases worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("Two Word Phrases")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Phrase", bold)

        # Add actual data
        i = 2
        for phrase, verse in sorted(self.twoWordPhrases.items()):
            worksheet.write("A" + str(i), verse[0])
            worksheet.write("B" + str(i), verse[1])
            worksheet.write("C" + str(i), verse[2])
            worksheet.write_rich_string("D" + str(i), *self.boldUniqueWords(phrase, bold))
            i += 1

        ################################################################################################################
        # Add Three Word Phrases worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("Three Word Phrases")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Phrase", bold)

        # Add actual data
        i = 2
        for phrase, verse in sorted(self.threeWordPhrases.items()):
            worksheet.write("A" + str(i), verse[0])
            worksheet.write("B" + str(i), verse[1])
            worksheet.write("C" + str(i), verse[2])
            worksheet.write_rich_string("D" + str(i), *self.boldUniqueWords(phrase, bold))
            i += 1

        ################################################################################################################
        # Add FTVs worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("FTVs")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Verse Start", bold)

        # Add actual data
        i = 2
        for verse in self.ftvs:
            worksheet.write("A" + str(i), verse[3])
            worksheet.write("B" + str(i), verse[4])
            worksheet.write("C" + str(i), verse[5])
            worksheet.write_rich_string("D" + str(i), *self.boldUniqueWords(verse[0], bold))
            i += 1
        ################################################################################################################
        # Add FTs worksheet
        ################################################################################################################
        worksheet = workbook.add_worksheet("FTs")

        # Add headers
        worksheet.write("A1", "Book", bold)
        worksheet.write("B1", "Chapter", bold)
        worksheet.write("C1", "Verse", bold)
        worksheet.write("D1", "Verse Start", bold)

        # Add actual data
        i = 2
        for verse in self.fts:
            worksheet.write("A" + str(i), verse[3])
            worksheet.write("B" + str(i), verse[4])
            worksheet.write("C" + str(i), verse[5])
            worksheet.write_rich_string("D" + str(i), *self.boldUniqueWords(verse[0], bold))
            i += 1

        try:
            workbook.close()  # Close the workbook
        except IOError:
            return "Error => Output file open!!!"

        return 0  # Return with no errors

    ####################################################################################################################
    # Helper Funcs
    ####################################################################################################################
    def splitVerse(self, verseText):
        """
        Function to split a verse into individual words.

        Parameters:
            verseText(str): Text of verse to be split.

        Returns:
            splitVerse(arr): Array of array of str and int representing split verse text.
        """
        splitVerse = []
        partOfWord = ["", 0]

        # Loop through all characters with index
        for i, character in enumerate(verseText):
            # Character is part of word
            if (character.isalnum()) or \
            (character == "-") or \
            (character in ["’" , "'"] and partOfWord[0] != "" and
            ((verseText[i - 5:i].lower() == "jesus") or
            (i < len(verseText) and verseText[i - 1].isalnum() and verseText[i + 1].isalnum()))):
                if partOfWord[0] == "":
                    partOfWord[1] = i
                partOfWord[0] += character

            # Character is not part of word
            else:
                if partOfWord[0]:
                    splitVerse.append(partOfWord.copy())
                    partOfWord[0] = ""
                    partOfWord[1] = 0

        # Append any leftover characters to splitVerse
        if partOfWord[0]:
            splitVerse.append(partOfWord.copy())

        return splitVerse

    def boldUniqueWords(self, myString, boldFormat):
        """
        Function to bold unique words in a particular string.

        Parameters:
            myString (str): The input string to be bolded.
            boldFormat (xlsxwriter format object): The format to be applied to unique words.

        Returns:
            result (array) Array of strings and xlsxwriter objects.
        """

        result = []
        start = 0
        splitVerse = self.splitVerse(myString)
        for word in splitVerse:
            if word[0].upper() in self.uniqueWords:
                result.append(myString[start:word[1]])
                result.append(boldFormat)
                result.append(myString[word[1]:word[1] + len(word[0])])
                start = len(word[0]) + word[1]
        if start != len(myString):
            result.append(myString[start:])
        return result


if __name__ == "__main__":
    start_time = time.time()

    lM = ListMaker()
    lM.importVerses()
    lM.createConcordance()
    lM.createUniqueWords()
    lM.createTwoWordPhrases()
    lM.createThreeWordPhrases()
    lM.createFtvs()
    lM.createFts()
    lM.exportLists()

    print("Done in: {:.2f}s".format(time.time() - start_time))
